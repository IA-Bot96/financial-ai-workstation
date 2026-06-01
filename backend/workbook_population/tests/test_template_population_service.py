"""Tests for populating compatible workbook templates."""

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

BACKEND_DIR = Path(__file__).resolve().parents[2]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from ocr_engine.models.insights_extraction import Insight
from shared.models.metric_value import MetricValue
from workbook_population.models.sheet_validation_result import SheetValidationResult
from workbook_population.services.template_population_service import (
    TemplatePopulationService,
)
from workbook_population.services.workbook_mapper import (
    WorkbookCellMapping,
    WorkbookMapper,
)


def _metric_value(
    metric: str,
    year: int,
    value: int,
    table_type: str = "income_statement",
) -> MetricValue:
    return MetricValue(
        metric=metric,
        value_year=year,
        value=value,
        source_report_year=2025,
        page_number=120,
        table_type=table_type,
    )


def test_template_population_preserves_formulas_and_writes_values(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Income Statement"
    worksheet.append(["Metric", 2024, 2025])
    worksheet.append(["revenue", None, None])
    worksheet.append(["gross_profit", None, "=C2*0.3"])
    workbook.save(template_path)
    workbook.close()

    sheets_reused, sheets_replaced, sheets_created, metrics_written, warnings = (
        TemplatePopulationService().populate(
            template_path=str(template_path),
            output_file_path=str(output_path),
            metric_values=[
                _metric_value("revenue", 2024, 1000),
                _metric_value("gross_profit", 2025, 300),
            ],
            insights=[
                Insight(
                    value_year=2025,
                    source_report_year=2025,
                    area="Debt",
                    takeaway="Borrowings increased.",
                    source_section="Business Review",
                    page_number=84,
                    confidence=0.9,
                )
            ],
            sheet_results=[
                SheetValidationResult(
                    sheet_name="Income Statement",
                    match_score=100,
                    is_compatible=True,
                    missing_metrics=[],
                    extra_metrics=[],
                    warnings=[],
                )
            ],
        )
    )

    populated = load_workbook(output_path, data_only=False)
    assert populated["Income Statement"]["B2"].value == 1000
    assert populated["Income Statement"]["C3"].value == "=C2*0.3"
    assert populated["Insights"]["A2"].value == 2025
    assert metrics_written == 1
    assert any("Skipped formula cell" in warning for warning in warnings)
    assert sheets_reused == ["Income Statement"]
    assert sheets_replaced == []
    assert "Insights" in sheets_created
    populated.close()


def test_template_population_replaces_incompatible_and_creates_missing_sheets(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Cash Flow"
    worksheet.append(["Legacy", 2025])
    worksheet.append(["old_metric", 1])
    workbook.save(template_path)
    workbook.close()

    sheets_reused, sheets_replaced, sheets_created, metrics_written, warnings = (
        TemplatePopulationService().populate(
            template_path=str(template_path),
            output_file_path=str(output_path),
            metric_values=[
                _metric_value(
                    "operating_cash_flow",
                    2025,
                    100,
                    table_type="cash_flow_statement",
                ),
                _metric_value(
                    "long_term_debt",
                    2025,
                    200,
                    table_type="debt_schedule",
                ),
            ],
            insights=[],
            sheet_results=[
                SheetValidationResult(
                    sheet_name="Cash Flow",
                    match_score=40,
                    is_compatible=False,
                    missing_metrics=["operating_cash_flow"],
                    extra_metrics=["old_metric"],
                    warnings=[],
                ),
                SheetValidationResult(
                    sheet_name="Debt Schedule",
                    match_score=0,
                    is_compatible=False,
                    missing_metrics=["long_term_debt"],
                    extra_metrics=[],
                    warnings=[],
                ),
            ],
        )
    )

    populated = load_workbook(output_path, data_only=False)
    assert populated["Cash Flow"]["A2"].value == "operating_cash_flow"
    assert populated["Debt Schedule"]["A2"].value == "long_term_debt"
    assert sheets_reused == []
    assert sheets_replaced == ["Cash Flow"]
    assert sheets_created == ["Debt Schedule", "Insights"]
    assert metrics_written == 2
    assert warnings == []
    populated.close()


def test_template_population_sanitizes_created_and_replaced_sheet_names(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Legacy"
    worksheet.append(["Metric", 2025])
    workbook.save(template_path)
    workbook.close()

    sheets_reused, sheets_replaced, sheets_created, metrics_written, warnings = (
        TemplatePopulationService().populate(
            template_path=str(template_path),
            output_file_path=str(output_path),
            metric_values=[
                _metric_value(
                    "commitments",
                    2025,
                    100,
                    table_type="contingent_liabilities_and_assets_note",
                ),
            ],
            insights=[],
            sheet_results=[
                SheetValidationResult(
                    sheet_name="Contingent Liabilities And Assets Note",
                    match_score=0,
                    is_compatible=False,
                    missing_metrics=["commitments"],
                    extra_metrics=[],
                    warnings=[],
                )
            ],
        )
    )

    populated = load_workbook(output_path, data_only=False)
    assert "Contingent Liabilities And Ass" in populated.sheetnames
    assert all(len(sheet_name) <= 31 for sheet_name in populated.sheetnames)
    assert populated["Contingent Liabilities And Ass"]["A2"].value == "commitments"
    assert sheets_reused == []
    assert sheets_replaced == []
    assert "Contingent Liabilities And Ass" in sheets_created
    assert metrics_written == 1
    assert warnings == []
    populated.close()


def test_template_population_replaces_cross_sheet_mapping_conflicts(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Income Statement"
    worksheet.append(["Metric", 2025])
    worksheet.append(["revenue", None])
    balance_sheet = workbook.create_sheet("Balance Sheet")
    balance_sheet.append(["Metric", 2025])
    balance_sheet.append(["revenue", None])
    workbook.save(template_path)
    workbook.close()

    service = TemplatePopulationService(
        mapper=WorkbookMapper(
            explicit_mappings={
                ("revenue", 2025, "income_statement"): WorkbookCellMapping(
                    "Balance Sheet",
                    2,
                    2,
                )
            }
        )
    )

    sheets_reused, sheets_replaced, sheets_created, metrics_written, warnings = (
        service.populate(
            template_path=str(template_path),
            output_file_path=str(output_path),
            metric_values=[_metric_value("revenue", 2025, 1000)],
            insights=[],
            sheet_results=[
                SheetValidationResult(
                    sheet_name="Income Statement",
                    match_score=100,
                    is_compatible=True,
                    missing_metrics=[],
                    extra_metrics=[],
                    warnings=[],
                )
            ],
        )
    )

    populated = load_workbook(output_path, data_only=False)
    assert populated["Income Statement"]["A2"].value == "revenue"
    assert populated["Income Statement"]["B2"].value == 1000
    assert sheets_reused == []
    assert sheets_replaced == ["Income Statement"]
    assert "Insights" in sheets_created
    assert metrics_written == 1
    assert any("Cross-sheet mapping conflict" in warning for warning in warnings)
    populated.close()


def test_template_population_routes_review_insights_without_exporting_rejects(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Income Statement"
    worksheet.append(["Metric", 2025])
    worksheet.append(["revenue", None])
    workbook.save(template_path)
    workbook.close()

    sheets_reused, sheets_replaced, sheets_created, metrics_written, warnings = (
        TemplatePopulationService().populate(
            template_path=str(template_path),
            output_file_path=str(output_path),
            metric_values=[_metric_value("revenue", 2025, 1000)],
            insights=[
                Insight(
                    value_year=2025,
                    source_report_year=2025,
                    area="Exports",
                    takeaway="Export sales increased by 20%.",
                    source_section="Business Review",
                    page_number=84,
                    confidence=0.9,
                ),
                Insight(
                    value_year=2025,
                    source_report_year=2025,
                    area="Exports",
                    takeaway="Export earnings are tracked as a growth lever.",
                    source_section="Business Review",
                    page_number=85,
                    confidence=0.6,
                ),
                Insight(
                    value_year=2025,
                    source_report_year=2025,
                    area="Internal controls",
                    takeaway="Adequate internal financial controls are in place.",
                    source_section="Risks",
                    page_number=86,
                    confidence=0.6,
                ),
            ],
            sheet_results=[
                SheetValidationResult(
                    sheet_name="Income Statement",
                    match_score=100,
                    is_compatible=True,
                    missing_metrics=[],
                    extra_metrics=[],
                    warnings=[],
                )
            ],
        )
    )

    populated = load_workbook(output_path, data_only=False)
    assert populated["Insights"]["C2"].value == "Exports"
    assert populated["Insights"].max_row == 2
    assert populated["Insights Review"]["C2"].value == "Exports"
    assert populated["Insights Review"].max_row == 2
    assert sheets_reused == ["Income Statement"]
    assert sheets_replaced == []
    assert sheets_created == ["Insights", "Insights Review"]
    assert metrics_written == 1
    assert warnings == []
    populated.close()
