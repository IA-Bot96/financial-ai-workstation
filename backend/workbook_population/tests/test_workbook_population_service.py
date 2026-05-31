"""Integration-style tests for workbook population orchestration."""

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

BACKEND_DIR = Path(__file__).resolve().parents[2]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from ocr_engine.models.insights_extraction import Insight
from ocr_engine.models.table_normalization import NormalizationResult
from shared.models.company_context import CompanyContext
from shared.models.metric_value import MetricValue
from shared.models.report import Report
from workbook_population.services.workbook_population_service import (
    OpenPyXLWorkbookPopulationService,
)


def _metric_value(
    metric: str,
    year: int,
    value: int,
    table_type: str = "income_statement",
) -> MetricValue:
    return MetricValue(
        metric=metric,
        value_year=year,
        value=value,
        source_report_year=2025,
        page_number=120,
        table_type=table_type,
    )


def _insight() -> Insight:
    return Insight(
        value_year=2024,
        source_report_year=2025,
        area="Debt",
        takeaway="Borrowings increased.",
        source_section="Business Review",
        page_number=84,
        confidence=0.9,
    )


def _report(year: int) -> Report:
    return Report(
        id=f"rpt_{year}",
        company_name="Maple Leaf Cement Factory Limited",
        year=year,
        file_name=f"MLCF_{year}_Annual_Report.pdf",
        file_path=f"/reports/MLCF_{year}_Annual_Report.pdf",
    )


def _save_template(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Income Statement"
    worksheet.append(["Metric", 2024, 2025])
    for metric in ["revenue", "gross_profit", "ebitda", "profit_after_tax"]:
        worksheet.append([metric, None, None])
    workbook.create_sheet("Balance Sheet")
    workbook.create_sheet("Cash Flow")
    workbook.save(path)
    workbook.close()


def test_workbook_population_uses_high_match_template(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    _save_template(template_path)

    service = OpenPyXLWorkbookPopulationService(
        output_dir=tmp_path,
        output_file_name="model.xlsx",
    )
    result = service.generate_workbook(
        metric_values=[
            _metric_value("revenue", 2024, 1000),
            _metric_value("gross_profit", 2025, 400),
            _metric_value("ebitda", 2025, 300),
            _metric_value("profit_after_tax", 2024, 180),
        ],
        insights=[_insight()],
        template_path=str(template_path),
    )

    workbook = load_workbook(result.output_file_path)
    assert result.workbook_mode == "template"
    assert result.workbook_match_score is not None
    assert result.workbook_match_score >= 95
    assert result.sheets_reused == ["Income Statement"]
    assert result.metrics_written == 4
    assert workbook["Income Statement"]["B2"].value == 1000
    workbook.close()


def test_workbook_population_generates_dynamic_without_template(tmp_path: Path) -> None:
    service = OpenPyXLWorkbookPopulationService(
        output_dir=tmp_path,
        output_file_name="dynamic.xlsx",
    )

    result = service.generate_workbook(
        metric_values=[_metric_value("revenue", 2024, 1000)],
        insights=[_insight()],
        template_path=None,
    )

    assert result.workbook_mode == "dynamic"
    assert result.workbook_match_score == 0
    assert result.metrics_written == 1
    assert Path(result.output_file_path).exists()
    assert Path(result.output_file_path).name.startswith("dynamic_")
    assert Path(result.output_file_path).suffix == ".xlsx"


def test_workbook_population_rejects_duplicate_metric_years(tmp_path: Path) -> None:
    service = OpenPyXLWorkbookPopulationService(output_dir=tmp_path)

    with pytest.raises(ValueError, match="Duplicate consolidated metric"):
        service.generate_workbook(
            metric_values=[
                _metric_value("revenue", 2024, 1000),
                _metric_value("revenue", 2024, 1100),
            ],
            insights=[],
            template_path=None,
        )


def test_workbook_population_allows_duplicate_metrics_across_table_types(
    tmp_path: Path,
) -> None:
    service = OpenPyXLWorkbookPopulationService(output_dir=tmp_path)

    result = service.generate_workbook(
        metric_values=[
            _metric_value("revenue", 2024, 1000, "income_statement"),
            _metric_value("revenue", 2024, 300, "segment_information"),
        ],
        insights=[],
        template_path=None,
    )

    workbook = load_workbook(result.output_file_path)
    assert workbook["Income Statement"]["B2"].value == 1000
    assert workbook["Segment Information"]["B2"].value == 300
    assert result.metrics_written == 2
    workbook.close()


def test_workbook_population_process_sets_generated_workbook(tmp_path: Path) -> None:
    service = OpenPyXLWorkbookPopulationService(
        output_dir=tmp_path,
        output_file_name="context_model.xlsx",
    )
    context = CompanyContext(
        company_name="Maple Leaf Cement Factory Limited",
        reports=[_report(2025)],
        normalization_results={2025: NormalizationResult(tables=[])},
        metric_values=[_metric_value("revenue", 2025, 1000)],
    )

    result = service.process(context)

    assert result is context
    assert context.generated_workbook is not None
    assert context.workbook_result == context.generated_workbook
    generated_path = Path(context.generated_workbook.output_file_path)
    assert generated_path.name.startswith("context_model_")
    assert generated_path.suffix == ".xlsx"
