"""Tests for OCR V2 Bridge Phase B3 real Lucky end-to-end run."""

from __future__ import annotations

import json
import sys
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

BACKEND_DIR = Path(__file__).resolve().parents[2]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from ocr import (  # noqa: E402
    OCRV2LuckyRun,
    WORKBOOK_HEADERS,
    WORKBOOK_SHEET_NAME,
    write_ocr_v2_b3_report,
)


def _workspace_tmp(test_name: str) -> Path:
    path = Path("output/ocr_v2_test_artifacts") / f"{test_name}_{uuid4().hex}"
    path.mkdir(parents=True, exist_ok=True)
    return path


def test_lucky_b3_run_executes_real_pipeline_without_oracle_input() -> None:
    tmp_path = _workspace_tmp("lucky_b3_run")
    workbook_path = tmp_path / "ocr_v2_lucky_workbook.xlsx"

    result = OCRV2LuckyRun().run(workbook_path=workbook_path)
    audit = result.audit

    assert audit.real_extraction_run is True
    assert audit.oracle_injected_values is False
    assert audit.tables_processed == 27
    assert audit.candidate_rows_generated > 0
    assert audit.registry_candidates > 0
    assert audit.canonical_values_selected > 0
    assert audit.multi_candidate_metric_year_groups > 0
    assert audit.governance_executed is True
    assert audit.selection_executed is True
    assert audit.workbook_generated is True
    assert audit.provenance_preserved is True
    assert audit.integrity_violations == ()


def test_lucky_b3_workbook_contains_selected_canonical_rows() -> None:
    tmp_path = _workspace_tmp("lucky_b3_workbook")
    workbook_path = tmp_path / "ocr_v2_lucky_workbook.xlsx"

    result = OCRV2LuckyRun().run(workbook_path=workbook_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[WORKBOOK_SHEET_NAME]

    assert workbook_path.exists()
    assert result.workbook_output.workbook_rows_generated == (
        result.audit.canonical_values_selected
    )
    assert [cell.value for cell in worksheet[1]] == list(WORKBOOK_HEADERS)
    assert worksheet.max_row == result.audit.canonical_values_selected + 1


def test_lucky_b3_artifacts_are_written_with_required_success_values() -> None:
    tmp_path = _workspace_tmp("lucky_b3_artifacts")
    workbook_path = tmp_path / "ocr_v2_lucky_workbook.xlsx"
    candidates_path = tmp_path / "ocr_v2_lucky_candidates.json"
    registry_path = tmp_path / "ocr_v2_lucky_registry.json"
    audit_path = tmp_path / "ocr_v2_lucky_run_audit.json"
    report_path = tmp_path / "ocr_v2_b3_report.json"

    report = write_ocr_v2_b3_report(
        workbook_path=workbook_path,
        candidates_path=candidates_path,
        registry_path=registry_path,
        audit_path=audit_path,
        report_path=report_path,
    )
    audit_payload = json.loads(audit_path.read_text(encoding="utf-8"))
    report_payload = json.loads(report_path.read_text(encoding="utf-8"))
    candidates_payload = json.loads(candidates_path.read_text(encoding="utf-8"))
    registry_payload = json.loads(registry_path.read_text(encoding="utf-8"))

    assert workbook_path.exists()
    assert candidates_path.exists()
    assert registry_path.exists()
    assert audit_payload["candidate_rows_generated"] > 0
    assert audit_payload["registry_candidates"] > 0
    assert audit_payload["canonical_values_selected"] > 0
    assert audit_payload["integrity_violations"] == []
    assert audit_payload["real_extraction_run"] is True
    assert audit_payload["oracle_injected_values"] is False
    assert report_payload["phase"] == "B3"
    assert report_payload["scope"] == "real_lucky_end_to_end_run"
    assert report_payload["integrity_audit_passed"] is True
    assert report_payload["extraction_engine_changes_added"] is False
    assert report_payload["governance_changes_added"] is False
    assert report_payload["selection_changes_added"] is False
    assert report_payload["workbook_generation_changes_added"] is False
    assert report_payload["msil_export_changes_added"] is False
    assert report_payload["ranking_logic_added"] is False
    assert report_payload["scoring_logic_added"] is False
    assert report_payload["llm_logic_added"] is False
    assert candidates_payload["real_extraction_run"] is True
    assert candidates_payload["oracle_injected_values"] is False
    assert registry_payload["real_extraction_run"] is True
    assert registry_payload["oracle_injected_values"] is False
    assert report.integrity_violations == ()
