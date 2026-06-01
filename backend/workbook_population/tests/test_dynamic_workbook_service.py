"""Tests for dynamic workbook generation."""

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl import load_workbook

BACKEND_DIR = Path(__file__).resolve().parents[2]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from ocr_engine.models.insights_extraction import Insight
from shared.models.metric_value import MetricValue
from workbook_population.services.dynamic_workbook_service import DynamicWorkbookService


def test_dynamic_workbook_service_creates_statement_and_insights_sheets(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "dynamic.xlsx"

    sheets, metrics_written, warnings = DynamicWorkbookService().generate(
        output_file_path=str(output_path),
        metric_values=[
            MetricValue(
                metric="revenue",
                value_year=2024,
                value=1500,
                source_report_year=2025,
                page_number=120,
                table_type="income_statement",
            ),
            MetricValue(
                metric="cash",
                value_year=2024,
                value=500,
                source_report_year=2025,
                page_number=121,
                table_type="balance_sheet",
            ),
            MetricValue(
                metric="long_term_debt",
                value_year=2024,
                value=300,
                source_report_year=2025,
                page_number=122,
                table_type="debt_schedule",
            ),
        ],
        insights=[
            Insight(
                value_year=2024,
                source_report_year=2025,
                area="Debt",
                takeaway="Borrowings increased.",
                source_section="Business Review",
                page_number=84,
                confidence=0.9,
            )
        ],
    )

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == [
        "Income Statement",
        "Balance Sheet",
        "Debt Schedule",
        "Insights",
    ]
    assert workbook["Income Statement"]["A2"].value == "revenue"
    assert workbook["Income Statement"]["B2"].value == 1500
    assert workbook["Balance Sheet"]["A2"].value == "cash"
    assert workbook["Debt Schedule"]["A2"].value == "long_term_debt"
    assert workbook["Insights"]["C2"].value == "Debt"
    assert metrics_written == 3
    assert warnings == []
    assert sheets == workbook.sheetnames
    workbook.close()


def test_dynamic_workbook_preserves_duplicate_metrics_across_table_types(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "dynamic.xlsx"

    sheets, metrics_written, warnings = DynamicWorkbookService().generate(
        output_file_path=str(output_path),
        metric_values=[
            MetricValue(
                metric="revenue",
                value_year=2024,
                value=1500,
                source_report_year=2025,
                page_number=120,
                table_type="income_statement",
            ),
            MetricValue(
                metric="revenue",
                value_year=2024,
                value=220,
                source_report_year=2025,
                page_number=121,
                table_type="segment_information",
            ),
        ],
        insights=[],
    )

    workbook = load_workbook(output_path)
    assert workbook["Income Statement"]["A2"].value == "revenue"
    assert workbook["Income Statement"]["B2"].value == 1500
    assert workbook["Segment Information"]["A2"].value == "revenue"
    assert workbook["Segment Information"]["B2"].value == 220
    assert metrics_written == 2
    assert warnings == []
    assert sheets == ["Income Statement", "Segment Information"]
    workbook.close()


def test_dynamic_workbook_sanitizes_generated_sheet_names(tmp_path: Path) -> None:
    output_path = tmp_path / "dynamic.xlsx"

    sheets, metrics_written, warnings = DynamicWorkbookService().generate(
        output_file_path=str(output_path),
        metric_values=[
            MetricValue(
                metric="commitments",
                value_year=2025,
                value=100,
                source_report_year=2025,
                page_number=120,
                table_type="contingent_liabilities_and_assets_note",
            ),
            MetricValue(
                metric="cash",
                value_year=2025,
                value=200,
                source_report_year=2025,
                page_number=121,
                table_type=r"cash/bank:*?[notes]",
            ),
        ],
        insights=[],
    )

    workbook = load_workbook(output_path)
    assert "Contingent Liabilities And Ass" in workbook.sheetnames
    assert "CashBankNotes" in workbook.sheetnames
    assert all(len(sheet_name) <= 31 for sheet_name in workbook.sheetnames)
    assert sheets == [
        "Contingent Liabilities And Ass",
        "CashBankNotes",
    ]
    assert metrics_written == 2
    assert warnings == []
    workbook.close()


def test_dynamic_workbook_routes_insights_by_confidence_governance(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "dynamic.xlsx"

    sheets, metrics_written, warnings = DynamicWorkbookService().generate(
        output_file_path=str(output_path),
        metric_values=[],
        insights=[
            Insight(
                value_year=2025,
                source_report_year=2025,
                area="Exports",
                takeaway="Export sales increased by 20%.",
                source_section="Business Review",
                page_number=84,
                confidence=0.9,
            ),
            Insight(
                value_year=2025,
                source_report_year=2025,
                area="Exports",
                takeaway="Export earnings are tracked as a growth lever.",
                source_section="Business Review",
                page_number=85,
                confidence=0.6,
            ),
            Insight(
                value_year=2025,
                source_report_year=2025,
                area="Internal controls",
                takeaway="Adequate internal financial controls are in place.",
                source_section="Risks",
                page_number=86,
                confidence=0.6,
            ),
            Insight(
                value_year=2025,
                source_report_year=2025,
                area="Sustainability",
                takeaway="Generated 799,551 kWh renewable electricity.",
                source_section="Sustainability",
                page_number=87,
                confidence=0.0,
            ),
        ],
    )

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["Insights", "Insights Review"]
    assert workbook["Insights"].max_row == 2
    assert workbook["Insights"]["C2"].value == "Exports"
    assert workbook["Insights Review"].max_row == 3
    assert workbook["Insights Review"]["C2"].value == "Exports"
    assert workbook["Insights Review"]["C3"].value == "Sustainability"
    assert "Internal controls" not in [
        workbook["Insights Review"][f"C{row}"].value
        for row in range(2, workbook["Insights Review"].max_row + 1)
    ]
    assert metrics_written == 0
    assert sheets == ["Insights", "Insights Review"]
    assert warnings == [
        "No metric values were provided; generated an empty model."
    ]
    workbook.close()


def test_dynamic_workbook_reports_permission_save_failure(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def raise_permission_error(self: Workbook, filename: object) -> None:
        raise PermissionError("locked")

    monkeypatch.setattr(Workbook, "save", raise_permission_error)

    with pytest.raises(PermissionError, match="Close the Excel file"):
        DynamicWorkbookService().generate(
            output_file_path=str(tmp_path / "locked.xlsx"),
            metric_values=[],
            insights=[],
        )
