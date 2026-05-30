"""Unit tests for workbook template compatibility scoring."""

import sys
from pathlib import Path

from openpyxl import Workbook

BACKEND_DIR = Path(__file__).resolve().parents[2]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from shared.models.metric_value import MetricValue
from workbook_population.services.template_structure_validator import (
    TemplateStructureValidator,
    _average,
)


def _metric_value(metric: str, year: int, table_type: str) -> MetricValue:
    return MetricValue(
        metric=metric,
        value_year=year,
        value=100,
        source_report_year=2025,
        page_number=120,
        table_type=table_type,
    )


def _save_template(path: Path, metrics: list[str], years: list[int]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Income Statement"
    worksheet.append(["Metric", *years])
    for metric in metrics:
        worksheet.append([metric, *([None] * len(years))])
    workbook.create_sheet("Balance Sheet")
    workbook.create_sheet("Cash Flow")
    workbook.save(path)
    workbook.close()


def test_template_validator_scores_high_match(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    _save_template(
        template_path,
        ["revenue", "gross_profit", "ebitda", "profit_after_tax"],
        [2024, 2025],
    )

    result = TemplateStructureValidator().validate(
        str(template_path),
        [
            _metric_value("revenue", 2024, "income_statement"),
            _metric_value("gross_profit", 2025, "income_statement"),
            _metric_value("ebitda", 2025, "income_statement"),
            _metric_value("profit_after_tax", 2024, "income_statement"),
        ],
    )

    assert result.match_score >= 95
    assert result.is_match is True
    assert result.sheet_results[0].sheet_name == "Income Statement"
    assert result.sheet_results[0].is_compatible is True
    assert result.missing_metrics == []


def test_template_validator_reports_missing_metrics_and_years(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    _save_template(template_path, ["revenue"], [2025])

    result = TemplateStructureValidator().validate(
        str(template_path),
        [
            _metric_value("revenue", 2025, "income_statement"),
            _metric_value("ebitda", 2024, "income_statement"),
        ],
    )

    assert result.match_score < 95
    assert result.sheet_results[0].match_score < 95
    assert "ebitda" in result.missing_metrics
    assert any("year columns" in warning for warning in result.warnings)


def test_template_validator_marks_missing_sheet_for_creation(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    _save_template(template_path, ["revenue"], [2025])

    result = TemplateStructureValidator().validate(
        str(template_path),
        [_metric_value("cash", 2025, "balance_sheet")],
    )

    assert result.is_match is False
    assert result.sheet_results[0].sheet_name == "Balance Sheet"
    assert result.sheet_results[0].match_score == 0
    assert result.sheet_results[0].missing_metrics == ["cash"]
    assert any("missing and will be created" in warning for warning in result.warnings)


def test_workbook_match_score_is_average_of_sheet_scores() -> None:
    assert _average([98, 95, 71]) == 88
