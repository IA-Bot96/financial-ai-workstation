"""Workbook cell mapping helpers."""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any, Mapping

from openpyxl.worksheet.worksheet import Worksheet

from shared.models.metric_value import MetricValue
from workbook_population.constants.workbook_constants import (
    STATEMENT_SHEET_BY_TABLE_TYPE,
)


@dataclass(frozen=True)
class WorkbookCellMapping:
    """Resolved workbook destination for one metric value."""

    sheet_name: str
    row: int
    column: int


class WorkbookMapper:
    """Map canonical metric values to workbook cells.

    Explicit mappings can be injected for template-specific layouts. When no
    explicit mapping exists, the mapper discovers rows by metric labels and
    columns by year headers.
    """

    def __init__(
        self,
        explicit_mappings: Mapping[tuple[str, int, str], WorkbookCellMapping]
        | None = None,
    ) -> None:
        """Initialize the mapper with optional configured cell mappings."""

        self._explicit_mappings = dict(explicit_mappings or {})

    def resolve_template_mapping(
        self,
        workbook: Any,
        metric_value: MetricValue,
    ) -> WorkbookCellMapping | None:
        """Resolve the destination cell for a metric value in an existing workbook."""

        explicit = self._explicit_mappings.get(
            (
                metric_value.metric,
                metric_value.value_year,
                metric_value.table_type,
            )
        )
        if explicit is not None:
            return explicit

        candidate_sheets = self._candidate_sheets(workbook, metric_value)
        for worksheet in candidate_sheets:
            metric_row = self._find_metric_row(worksheet, metric_value.metric)
            year_column = self._find_year_column(worksheet, metric_value.value_year)
            if metric_row is not None and year_column is not None:
                return WorkbookCellMapping(
                    sheet_name=worksheet.title,
                    row=metric_row,
                    column=year_column,
                )

        return None

    def sheet_name_for_metric_value(self, metric_value: MetricValue) -> str:
        """Return the preferred dynamic workbook sheet for a metric value."""

        normalized_table_type = _normalize_key(metric_value.table_type)
        return STATEMENT_SHEET_BY_TABLE_TYPE.get(
            normalized_table_type,
            self._fallback_sheet_name(metric_value),
        )

    @staticmethod
    def _candidate_sheets(workbook: Any, metric_value: MetricValue) -> list[Worksheet]:
        preferred_name = STATEMENT_SHEET_BY_TABLE_TYPE.get(
            _normalize_key(metric_value.table_type)
        )
        if preferred_name in workbook.sheetnames:
            return [workbook[preferred_name]]
        if preferred_name is not None:
            normalized_preferred = _normalize_key(preferred_name)
            for sheet_name in workbook.sheetnames:
                if _normalize_key(sheet_name) == normalized_preferred:
                    return [workbook[sheet_name]]
        return list(workbook.worksheets)

    @staticmethod
    def _find_metric_row(worksheet: Worksheet, metric: str) -> int | None:
        target = _normalize_key(metric)
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and _normalize_key(cell.value) == target:
                    return cell.row
        return None

    @staticmethod
    def _find_year_column(worksheet: Worksheet, value_year: int) -> int | None:
        for row in worksheet.iter_rows(max_row=min(20, worksheet.max_row)):
            for cell in row:
                if _cell_year(cell.value) == value_year:
                    return cell.column
        return None

    @staticmethod
    def _fallback_sheet_name(metric_value: MetricValue) -> str:
        table_type = metric_value.table_type.replace("_", " ").strip().title()
        return table_type or "Financial Data"


def _normalize_key(value: str) -> str:
    """Normalize workbook labels to canonical snake-case-like keys."""

    normalized = str(value).lower().replace("&", " and ")
    normalized = re.sub(r"[^a-z0-9]+", "_", normalized)
    return normalized.strip("_")


def _cell_year(value: object) -> int | None:
    """Return a year from an Excel cell value when it is year-like."""

    if isinstance(value, int) and 1900 <= value <= 2200:
        return value
    if isinstance(value, float) and value.is_integer() and 1900 <= value <= 2200:
        return int(value)
    if isinstance(value, str):
        match = re.fullmatch(r"(?:19|20|21)\d{2}", value.strip())
        if match:
            return int(match.group(0))
    return None


__all__ = ["WorkbookCellMapping", "WorkbookMapper"]
