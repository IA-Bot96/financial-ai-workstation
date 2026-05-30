"""End-to-end OCR pipeline integration test.

This test intentionally mocks only OpenAI network calls. Table detection, table
extraction, metric normalization, validation, consolidation, and workbook
population all use their production implementations when the required fixture
PDF and OCR dependencies are available.
"""

from __future__ import annotations

import importlib.util
import json
import os
from pathlib import Path
import sys
import warnings
from typing import Any, Callable

import pytest

BACKEND_DIR = Path(__file__).resolve().parents[2] / "backend"
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

FIXTURE_PDF_PATH = Path(__file__).resolve().parents[1] / "fixtures" / "sample_annual_report.pdf"
E2E_TIMEOUT_WARNING_SECONDS = float(os.getenv("OCR_E2E_MAX_SECONDS", "300"))

pytestmark = pytest.mark.integration


class _FakeOpenAIResponse:
    """OpenAI response test double exposing the SDK's output_text shape."""

    def __init__(self, payload: dict[str, Any]) -> None:
        self.output_text = json.dumps(payload)


class _FakeResponses:
    """Responses API test double for deterministic OpenAI structured output."""

    def __init__(self, payload_factory: Callable[[dict[str, Any]], dict[str, Any]]) -> None:
        self._payload_factory = payload_factory
        self.calls: list[dict[str, Any]] = []

    def create(self, **kwargs: Any) -> _FakeOpenAIResponse:
        self.calls.append(kwargs)
        return _FakeOpenAIResponse(self._payload_factory(kwargs))


class _FakeOpenAIClient:
    """OpenAI client test double with a Responses API surface."""

    def __init__(self, payload_factory: Callable[[dict[str, Any]], dict[str, Any]]) -> None:
        self.responses = _FakeResponses(payload_factory)


def test_ocr_pipeline_processes_sample_annual_report_end_to_end(
    tmp_path: Path,
) -> None:
    """Process a real annual-report PDF into a generated Excel workbook."""

    if not FIXTURE_PDF_PATH.exists():
        pytest.skip(
            "Missing integration fixture: tests/fixtures/sample_annual_report.pdf. "
            "Place a real annual report PDF there to run this test."
        )

    missing_dependencies = _missing_dependencies(
        [
            "camelot",
            "fitz",
            "openpyxl",
            "pdfplumber",
            "PIL",
            "sentence_transformers",
            "torch",
            "transformers",
        ]
    )
    if missing_dependencies:
        pytest.skip(
            "Missing OCR integration dependencies: "
            f"{', '.join(missing_dependencies)}"
        )

    from openpyxl import load_workbook

    from ocr_engine.models.insights_extraction import Insight
    from ocr_engine.pipeline.models.pipeline_status import PipelineStatus
    from ocr_engine.pipeline.ocr_pipeline import OCRPipeline
    from ocr_engine.services.camelot_table_extractor import CamelotTableExtractor
    from ocr_engine.services.insights_extractor import InsightsExtractor
    from ocr_engine.services.openai_insights_extractor import OpenAIInsightsExtractor
    from ocr_engine.services.openai_table_classifier import OpenAITableClassifier
    from ocr_engine.services.table_metric_normalizer import TableMetricNormalizer
    from ocr_engine.services.table_transformer_detector import TableTransformerDetector
    from ocr_engine.validation.financial_validation_service import FinancialValidationService
    from shared.models.company_context import CompanyContext
    from shared.models.report import Report
    from shared.services.financial_year_consolidator import FinancialYearConsolidator
    from workbook_population.services.workbook_population_service import (
        OpenPyXLWorkbookPopulationService,
    )

    report_year = int(os.getenv("OCR_E2E_REPORT_YEAR", "2024"))
    context = CompanyContext(
        company_name=os.getenv("OCR_E2E_COMPANY_NAME", "Integration Test Company"),
        reports=[
            Report(
                id="e2e_sample_annual_report",
                company_name=os.getenv(
                    "OCR_E2E_COMPANY_NAME",
                    "Integration Test Company",
                ),
                year=report_year,
                file_name=FIXTURE_PDF_PATH.name,
                file_path=str(FIXTURE_PDF_PATH),
            )
        ],
    )

    classification_client = _FakeOpenAIClient(_classification_payload)
    insights_client = _FakeOpenAIClient(
        lambda _: {
            "insights": [
                Insight(
                    value_year=report_year,
                    source_report_year=report_year,
                    area="Operations",
                    takeaway="Management discussed operational performance in the report.",
                    source_section="Business Review",
                    page_number=1,
                    confidence=0.85,
                ).model_dump()
            ]
        }
    )

    pipeline = OCRPipeline(
        table_detector=TableTransformerDetector(),
        table_classifier=OpenAITableClassifier(
            client=classification_client,
            api_key="test-openai-key",
            model="gpt-5",
        ),
        table_extractor=CamelotTableExtractor(),
        validator=FinancialValidationService(),
        metric_normalizer=TableMetricNormalizer(),
        insights_extractor=OpenAIInsightsExtractor(
            insights_extractor=InsightsExtractor(
                client=insights_client,
                model="gpt-5",
                max_retries=1,
                retry_backoff_seconds=0,
            )
        ),
        financial_year_consolidator=FinancialYearConsolidator(),
        workbook_population_service=OpenPyXLWorkbookPopulationService(
            output_dir=tmp_path,
            output_file_name="ocr_e2e_output.xlsx",
        ),
    )

    result = pipeline.process(context)

    assert result.pipeline_status is PipelineStatus.COMPLETED
    assert result.pipeline_errors == []

    workbook_result = result.generated_workbook
    assert workbook_result is not None
    workbook_path = Path(workbook_result.output_file_path)
    assert workbook_path.exists()

    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        assert workbook.sheetnames
        workbook_values = [
            str(cell.value).strip().lower()
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        ]
    finally:
        workbook.close()

    detected_table_count = sum(
        detected_page.tables_detected
        for result_by_year in result.table_detection_results.values()
        for detected_page in result_by_year.detected_pages
    )
    extracted_table_count = sum(
        len(extraction_result.tables)
        for extraction_result in result.extraction_results.values()
    )
    normalized_metric_count = sum(
        len(normalization_result.metric_values)
        for normalization_result in result.normalization_results.values()
    )

    assert detected_table_count > 0
    assert extracted_table_count > 0
    assert result.metric_values
    assert normalized_metric_count > 0
    assert result.validation_results
    assert result.normalization_results
    assert any(metric.metric.lower() in workbook_values for metric in result.metric_values)

    assert result.execution_results
    assert all(
        execution_result.execution_time_seconds >= 0
        for execution_result in result.execution_results
    )
    total_execution_seconds = sum(
        execution_result.execution_time_seconds
        for execution_result in result.execution_results
    )
    if total_execution_seconds > E2E_TIMEOUT_WARNING_SECONDS:
        warnings.warn(
            "OCR E2E pipeline exceeded expected duration: "
            f"{total_execution_seconds:.2f}s > {E2E_TIMEOUT_WARNING_SECONDS:.2f}s",
            RuntimeWarning,
            stacklevel=2,
        )

    print(
        "\nOCR E2E Test Report\n"
        f"tables_detected: {detected_table_count}\n"
        f"tables_extracted: {extracted_table_count}\n"
        f"metrics_normalized: {normalized_metric_count}\n"
        f"consolidated_metric_values: {len(result.metric_values)}\n"
        f"workbook_path: {workbook_path}\n"
        f"execution_seconds: {total_execution_seconds:.2f}"
    )


def _classification_payload(kwargs: dict[str, Any]) -> dict[str, list[str]]:
    """Return deterministic table types without calling OpenAI."""

    prompt_text = " ".join(
        str(message.get("content", ""))
        for message in kwargs.get("input", [])
        if isinstance(message, dict)
    ).lower()

    if "cash flow" in prompt_text or "cash flows" in prompt_text:
        return {"table_types": ["cash_flow_statement"]}
    if "financial position" in prompt_text or "balance sheet" in prompt_text:
        return {"table_types": ["balance_sheet"]}
    if "debt" in prompt_text or "borrowings" in prompt_text:
        return {"table_types": ["debt_schedule"]}
    return {"table_types": ["income_statement"]}


def _missing_dependencies(module_names: list[str]) -> list[str]:
    """Return import names that are unavailable in the active test environment."""

    return [
        module_name
        for module_name in module_names
        if importlib.util.find_spec(module_name) is None
    ]
