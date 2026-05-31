"""Excel template compatibility scoring."""

from __future__ import annotations

import logging
from collections import defaultdict
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from shared.models.metric_value import MetricValue
from workbook_population.constants.workbook_constants import (
    CRITICAL_METRICS,
    HIGH_TEMPLATE_MATCH_THRESHOLD,
    LOW_TEMPLATE_MATCH_THRESHOLD,
)
from workbook_population.models.sheet_validation_result import SheetValidationResult
from workbook_population.models.template_validation_result import (
    TemplateValidationResult,
)
from workbook_population.services.workbook_mapper import (
    WorkbookMapper,
    _cell_year,
    _normalize_key,
)

logger = logging.getLogger(__name__)

_IGNORED_TEMPLATE_LABELS = {
    "account",
    "amount",
    "description",
    "line_item",
    "metric",
    "metrics",
    "note",
    "notes",
    "particular",
    "particulars",
    "pkr",
    "pkrs",
    "rs",
    "rupees",
    "year",
    "years",
}


class TemplateStructureValidator:
    """Score an Excel template against extracted financial metric values per sheet."""

    def __init__(
        self,
        *,
        auto_match_threshold: float = HIGH_TEMPLATE_MATCH_THRESHOLD,
        user_decision_threshold: float = LOW_TEMPLATE_MATCH_THRESHOLD,
        mapper: WorkbookMapper | None = None,
        log: logging.Logger | None = None,
    ) -> None:
        """Initialize template scoring rules."""

        if not 0 <= user_decision_threshold <= auto_match_threshold <= 100:
            raise ValueError(
                "Thresholds must satisfy 0 <= user_decision_threshold "
                "<= auto_match_threshold <= 100."
            )

        self._auto_match_threshold = auto_match_threshold
        self._user_decision_threshold = user_decision_threshold
        self._mapper = mapper or WorkbookMapper()
        self._logger = log or logger

    def validate(
        self,
        template_path: str,
        metric_values: list[MetricValue],
    ) -> TemplateValidationResult:
        """Validate template compatibility using independent sheet-level scoring."""

        if not metric_values:
            return TemplateValidationResult(
                is_match=False,
                match_score=0,
                sheet_results=[],
                missing_metrics=[],
                extra_metrics=[],
                warnings=["No metric values were provided for template validation."],
            )

        grouped_values = self._group_values_by_sheet(metric_values)

        try:
            workbook = load_workbook(template_path, data_only=False)
        except Exception as exc:
            self._logger.exception(
                "Failed to load workbook template",
                extra={"template_path": template_path},
            )
            return TemplateValidationResult(
                is_match=False,
                match_score=0,
                sheet_results=[],
                missing_metrics=sorted(_metric_set(metric_values)),
                extra_metrics=[],
                warnings=[f"Template could not be loaded: {exc}"],
            )

        try:
            sheet_results: list[SheetValidationResult] = []
            for sheet_name, values in grouped_values.items():
                worksheet = self._find_sheet(workbook, sheet_name)
                if worksheet is None:
                    sheet_results.append(self._missing_sheet_result(sheet_name, values))
                    continue

                sheet_results.append(
                    self._score_sheet(
                        sheet_name=sheet_name,
                        worksheet=worksheet,
                        metric_values=values,
                    )
                )

            match_score = _average(
                sheet_result.match_score for sheet_result in sheet_results
            )
            missing_metrics = sorted(
                {
                    metric
                    for sheet_result in sheet_results
                    for metric in sheet_result.missing_metrics
                }
            )
            extra_metrics = sorted(
                {
                    metric
                    for sheet_result in sheet_results
                    for metric in sheet_result.extra_metrics
                }
            )
            warnings = [
                warning
                for sheet_result in sheet_results
                for warning in sheet_result.warnings
            ]

            return TemplateValidationResult(
                is_match=bool(sheet_results)
                and all(sheet_result.is_compatible for sheet_result in sheet_results),
                match_score=match_score,
                sheet_results=sheet_results,
                missing_metrics=missing_metrics,
                extra_metrics=extra_metrics,
                warnings=warnings,
            )
        finally:
            workbook.close()

    def _group_values_by_sheet(
        self,
        metric_values: Iterable[MetricValue],
    ) -> dict[str, list[MetricValue]]:
        grouped: dict[str, list[MetricValue]] = defaultdict(list)
        for metric_value in metric_values:
            grouped[self._mapper.sheet_name_for_metric_value(metric_value)].append(
                metric_value
            )
        return dict(grouped)

    @staticmethod
    def _find_sheet(workbook: object, sheet_name: str) -> Worksheet | None:
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]

        normalized_target = _normalize_key(sheet_name)
        for existing_sheet_name in workbook.sheetnames:
            if _normalize_key(existing_sheet_name) == normalized_target:
                return workbook[existing_sheet_name]
        return None

    def _missing_sheet_result(
        self,
        sheet_name: str,
        metric_values: list[MetricValue],
    ) -> SheetValidationResult:
        return SheetValidationResult(
            sheet_name=sheet_name,
            match_score=0,
            is_compatible=False,
            missing_metrics=sorted(_metric_set(metric_values)),
            extra_metrics=[],
            warnings=[f"{sheet_name} sheet is missing and will be created."],
        )

    def _score_sheet(
        self,
        *,
        sheet_name: str,
        worksheet: Worksheet,
        metric_values: list[MetricValue],
    ) -> SheetValidationResult:
        required_metrics = _metric_set(metric_values)
        template_metrics = self._extract_sheet_metrics(worksheet)
        required_years = {metric_value.value_year for metric_value in metric_values}
        template_years = self._extract_sheet_years(worksheet)

        metric_coverage = _coverage(required_metrics, template_metrics)
        year_coverage = _coverage(required_years, template_years)
        position_score = self._position_consistency_score(worksheet, required_metrics)
        critical_required_metrics = required_metrics & CRITICAL_METRICS
        critical_score = _coverage(critical_required_metrics, template_metrics)

        match_score = round(
            (metric_coverage * 50)
            + (year_coverage * 25)
            + (position_score * 15)
            + (critical_score * 10),
            2,
        )
        missing_metrics = sorted(required_metrics - template_metrics)
        extra_metrics = sorted(template_metrics - required_metrics)

        return SheetValidationResult(
            sheet_name=sheet_name,
            match_score=match_score,
            is_compatible=match_score >= self._auto_match_threshold,
            missing_metrics=missing_metrics,
            extra_metrics=extra_metrics,
            warnings=self._warnings(
                sheet_name=sheet_name,
                match_score=match_score,
                missing_metrics=missing_metrics,
                required_years=required_years,
                template_years=template_years,
            ),
        )

    @staticmethod
    def _extract_sheet_metrics(worksheet: Worksheet) -> set[str]:
        metrics: set[str] = set()
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    key = _normalize_key(cell.value)
                    if key and key not in _IGNORED_TEMPLATE_LABELS:
                        metrics.add(key)
        return metrics

    @staticmethod
    def _extract_sheet_years(worksheet: Worksheet) -> set[int]:
        years: set[int] = set()
        for row in worksheet.iter_rows(max_row=min(20, worksheet.max_row)):
            for cell in row:
                year = _cell_year(cell.value)
                if year is not None:
                    years.add(year)
        return years

    @staticmethod
    def _position_consistency_score(
        worksheet: Worksheet,
        required_metrics: set[str],
    ) -> float:
        if not required_metrics:
            return 0.0

        found_metrics: set[str] = set()
        for row in worksheet.iter_rows(max_col=min(3, worksheet.max_column)):
            for cell in row:
                if isinstance(cell.value, str):
                    key = _normalize_key(cell.value)
                    if key in required_metrics:
                        found_metrics.add(key)

        return len(found_metrics) / len(required_metrics)

    def _warnings(
        self,
        *,
        sheet_name: str,
        match_score: float,
        missing_metrics: list[str],
        required_years: set[int],
        template_years: set[int],
    ) -> list[str]:
        warnings: list[str] = []

        missing_critical = sorted(set(missing_metrics) & CRITICAL_METRICS)
        if missing_critical:
            warnings.append(
                f"{sheet_name} sheet missing critical metrics: "
                f"{', '.join(missing_critical)}."
            )

        for metric in missing_metrics[:5]:
            warnings.append(f"{_display_metric(metric)} row missing.")

        remaining_missing = len(missing_metrics) - 5
        if remaining_missing > 0:
            warnings.append(
                f"{sheet_name} sheet has {remaining_missing} additional "
                "missing metrics."
            )

        missing_years = sorted(required_years - template_years)
        if missing_years:
            warnings.append(
                f"{sheet_name} sheet missing year columns: "
                f"{', '.join(str(year) for year in missing_years)}."
            )

        if self._user_decision_threshold <= match_score < self._auto_match_threshold:
            warnings.append(
                f"{sheet_name} template is {match_score}% compatible and "
                "requires user decision."
            )
        elif match_score < self._user_decision_threshold:
            warnings.append(
                f"{sheet_name} sheet is below 80% compatibility and will be replaced."
            )

        return warnings


def _metric_set(metric_values: Iterable[MetricValue]) -> set[str]:
    return {_normalize_key(metric_value.metric) for metric_value in metric_values}


def _coverage(required: set[object], available: set[object]) -> float:
    if not required:
        return 0.0
    return len(required & available) / len(required)


def _average(values: Iterable[float]) -> float:
    value_list = list(values)
    if not value_list:
        return 0
    return round(sum(value_list) / len(value_list), 2)


def _display_metric(metric: str) -> str:
    return metric.replace("_", " ").title()


__all__ = ["TemplateStructureValidator"]
