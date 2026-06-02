"""Template workbook population service."""

from __future__ import annotations

import logging
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ocr_engine.models.insights_extraction import Insight
from ocr_engine.governance.insight_confidence_governance import (
    InsightConfidenceGovernance,
)
from shared.models.metric_value import MetricValue
from workbook_population.constants.workbook_constants import (
    INSIGHTS_REVIEW_SHEET_NAME,
    INSIGHTS_SHEET_NAME,
)
from workbook_population.models.sheet_validation_result import SheetValidationResult
from workbook_population.models.workbook_cell_mapping import WorkbookCellMappingDraft
from workbook_population.services.sheet_name_sanitizer import sanitize_sheet_name
from workbook_population.services.workbook_mapper import WorkbookMapper, _normalize_key

logger = logging.getLogger(__name__)


class CrossSheetMappingConflict(RuntimeError):
    """Raised when a template mapping targets a different sheet than expected."""


class TemplatePopulationService:
    """Populate Excel templates using independent sheet-level decisions."""

    def __init__(
        self,
        *,
        mapper: WorkbookMapper | None = None,
        log: logging.Logger | None = None,
    ) -> None:
        """Initialize the service with an injectable workbook mapper."""

        self._mapper = mapper or WorkbookMapper()
        self._logger = log or logger
        self._insight_governance = InsightConfidenceGovernance()
        self._last_cell_mapping_drafts: list[WorkbookCellMappingDraft] = []

    @property
    def last_cell_mapping_drafts(self) -> list[WorkbookCellMappingDraft]:
        """Return workbook cell mappings captured by the most recent populate call."""

        return list(self._last_cell_mapping_drafts)

    def populate(
        self,
        *,
        template_path: str,
        output_file_path: str,
        metric_values: list[MetricValue],
        insights: list[Insight],
        sheet_results: list[SheetValidationResult],
    ) -> tuple[list[str], list[str], list[str], int, list[str]]:
        """Populate, replace, or create sheets based on sheet-level compatibility."""

        self._last_cell_mapping_drafts = []
        workbook = load_workbook(template_path, data_only=False)
        metrics_written = 0
        warnings: list[str] = []
        sheets_reused: list[str] = []
        sheets_replaced: list[str] = []
        sheets_created: list[str] = []

        grouped_values = self._group_values_by_sheet(metric_values)
        sheet_result_by_name = {
            _normalize_key(sheet_result.sheet_name): sheet_result
            for sheet_result in sheet_results
        }

        try:
            for sheet_name, values in grouped_values.items():
                sheet_result = sheet_result_by_name.get(_normalize_key(sheet_name))
                existing_sheet_name = self._find_existing_sheet_name(
                    workbook,
                    sheet_name,
                )

                if existing_sheet_name is None:
                    safe_sheet_name = sanitize_sheet_name(
                        sheet_name,
                        set(workbook.sheetnames),
                    )
                    worksheet = workbook.create_sheet(safe_sheet_name)
                    metrics_written += self._write_generated_metric_sheet(
                        worksheet,
                        values,
                    )
                    sheets_created.append(safe_sheet_name)
                    continue

                if sheet_result is None:
                    warnings.append(
                        f"{sheet_name} has no validation result; sheet left unchanged."
                    )
                    continue

                if sheet_result.is_compatible:
                    try:
                        written, sheet_warnings = self._populate_compatible_sheet(
                            workbook=workbook,
                            sheet_name=sheet_name,
                            metric_values=values,
                        )
                    except CrossSheetMappingConflict as exc:
                        warning = f"{exc} Replacing sheet to prevent data loss."
                        warnings.append(warning)
                        self._logger.warning(
                            "Replacing sheet due to cross-sheet mapping conflict",
                            extra={"sheet_name": sheet_name},
                            exc_info=True,
                        )
                        replacement_index = workbook.sheetnames.index(
                            existing_sheet_name
                        )
                        workbook.remove(workbook[existing_sheet_name])
                        safe_sheet_name = sanitize_sheet_name(
                            sheet_name,
                            set(workbook.sheetnames),
                        )
                        worksheet = workbook.create_sheet(
                            safe_sheet_name,
                            replacement_index,
                        )
                        metrics_written += self._write_generated_metric_sheet(
                            worksheet,
                            values,
                        )
                        sheets_replaced.append(safe_sheet_name)
                        continue

                    metrics_written += written
                    warnings.extend(sheet_warnings)
                    sheets_reused.append(existing_sheet_name)
                    continue

                if _requires_user_decision(sheet_result):
                    warnings.append(
                        f"{sheet_name} template is {sheet_result.match_score}% "
                        "compatible and requires user decision."
                    )
                    continue

                replacement_index = workbook.sheetnames.index(existing_sheet_name)
                workbook.remove(workbook[existing_sheet_name])
                safe_sheet_name = sanitize_sheet_name(
                    sheet_name,
                    set(workbook.sheetnames),
                )
                worksheet = workbook.create_sheet(safe_sheet_name, replacement_index)
                metrics_written += self._write_generated_metric_sheet(worksheet, values)
                sheets_replaced.append(safe_sheet_name)

            sheets_created.extend(self._populate_insights(workbook, insights))

            output_path = Path(output_file_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            try:
                workbook.save(output_path)
            except PermissionError as exc:
                raise PermissionError(_save_permission_message(output_path)) from exc
            return (
                sheets_reused,
                sheets_replaced,
                sheets_created,
                metrics_written,
                warnings,
            )
        finally:
            workbook.close()

    def _populate_compatible_sheet(
        self,
        *,
        workbook: object,
        sheet_name: str,
        metric_values: list[MetricValue],
    ) -> tuple[int, list[str]]:
        metrics_written = 0
        warnings: list[str] = []

        for metric_value in metric_values:
            mapping = self._mapper.resolve_template_mapping(workbook, metric_value)
            if mapping is None:
                warnings.append(
                    "No template cell mapping found for "
                    f"{metric_value.metric} {metric_value.value_year}."
                )
                continue

            if _normalize_key(mapping.sheet_name) != _normalize_key(sheet_name):
                message = (
                    "Cross-sheet mapping conflict for "
                    f"{metric_value.metric} {metric_value.value_year}: "
                    f"expected {sheet_name}, got {mapping.sheet_name}."
                )
                self._logger.warning(
                    "Cross-sheet template mapping conflict",
                    extra={
                        "metric": metric_value.metric,
                        "value_year": metric_value.value_year,
                        "expected_sheet": sheet_name,
                        "mapped_sheet": mapping.sheet_name,
                    },
                )
                raise CrossSheetMappingConflict(message)

            worksheet = workbook[mapping.sheet_name]
            cell = worksheet.cell(row=mapping.row, column=mapping.column)
            if self._is_formula_cell(cell.value):
                warnings.append(
                    "Skipped formula cell " f"{mapping.sheet_name}!{cell.coordinate}."
                )
                self._last_cell_mapping_drafts.append(
                    WorkbookCellMappingDraft(
                        metric=metric_value.metric,
                        value_year=metric_value.value_year,
                        source_report_year=metric_value.source_report_year,
                        table_type=metric_value.table_type,
                        sheet_name=mapping.sheet_name,
                        row=mapping.row,
                        column=mapping.column,
                        cell_reference=cell.coordinate,
                        write_status="skipped_formula",
                        written_value=metric_value.value,
                    )
                )
                continue

            cell.value = metric_value.value
            self._last_cell_mapping_drafts.append(
                WorkbookCellMappingDraft(
                    metric=metric_value.metric,
                    value_year=metric_value.value_year,
                    source_report_year=metric_value.source_report_year,
                    table_type=metric_value.table_type,
                    sheet_name=mapping.sheet_name,
                    row=mapping.row,
                    column=mapping.column,
                    cell_reference=cell.coordinate,
                    write_status="written",
                    written_value=metric_value.value,
                )
            )
            metrics_written += 1

        return metrics_written, warnings

    def _group_values_by_sheet(
        self,
        metric_values: Iterable[MetricValue],
    ) -> dict[str, list[MetricValue]]:
        grouped: dict[str, list[MetricValue]] = defaultdict(list)
        for metric_value in metric_values:
            grouped[self._mapper.sheet_name_for_metric_value(metric_value)].append(
                metric_value
            )
        return dict(grouped)

    @staticmethod
    def _find_existing_sheet_name(workbook: object, sheet_name: str) -> str | None:
        if sheet_name in workbook.sheetnames:
            return sheet_name

        normalized_target = _normalize_key(sheet_name)
        sanitized_target = sanitize_sheet_name(sheet_name, set())
        normalized_sanitized_target = _normalize_key(sanitized_target)
        for existing_sheet_name in workbook.sheetnames:
            normalized_existing = _normalize_key(existing_sheet_name)
            if normalized_existing in {
                normalized_target,
                normalized_sanitized_target,
            }:
                return existing_sheet_name
        return None

    def _write_generated_metric_sheet(
        self,
        worksheet: Worksheet,
        metric_values: list[MetricValue],
    ) -> int:
        years = sorted({metric_value.value_year for metric_value in metric_values})
        row_keys = _ordered_unique(
            (metric_value.metric, metric_value.table_type)
            for metric_value in metric_values
        )
        duplicate_metrics = _duplicate_metrics(metric_values)
        values_by_metric_year = {
            (
                metric_value.metric,
                metric_value.value_year,
                metric_value.table_type,
            ): metric_value.value
            for metric_value in metric_values
        }
        metric_value_by_key = {
            (
                metric_value.metric,
                metric_value.value_year,
                metric_value.table_type,
            ): metric_value
            for metric_value in metric_values
        }

        worksheet.append(["Metric", *years])
        written = 0
        for metric, table_type in row_keys:
            row = [_row_label(metric, table_type, duplicate_metrics)]
            for year in years:
                value = values_by_metric_year.get((metric, year, table_type))
                row.append(value)
                if value is not None:
                    written += 1
            worksheet.append(row)
            row_number = worksheet.max_row
            for year_index, year in enumerate(years, start=2):
                metric_value = metric_value_by_key.get((metric, year, table_type))
                if metric_value is None or metric_value.value is None:
                    continue
                self._last_cell_mapping_drafts.append(
                    WorkbookCellMappingDraft(
                        metric=metric_value.metric,
                        value_year=metric_value.value_year,
                        source_report_year=metric_value.source_report_year,
                        table_type=metric_value.table_type,
                        sheet_name=worksheet.title,
                        row=row_number,
                        column=year_index,
                        cell_reference=f"{get_column_letter(year_index)}{row_number}",
                        write_status="written",
                        written_value=metric_value.value,
                    )
                )

        _autosize_columns(worksheet)
        return written

    def _populate_insights(self, workbook: object, insights: list[Insight]) -> list[str]:
        """Write governed insights without touching formulas elsewhere."""

        governance_result = self._insight_governance.apply(insights)
        sheets_created: list[str] = []
        created_sheet = self._write_insights_sheet(
            workbook,
            INSIGHTS_SHEET_NAME,
            governance_result.exported_insights,
        )
        if created_sheet is not None:
            sheets_created.append(created_sheet)
        if governance_result.review_insights:
            created_sheet = self._write_insights_sheet(
                workbook,
                INSIGHTS_REVIEW_SHEET_NAME,
                governance_result.review_insights,
            )
            if created_sheet is not None:
                sheets_created.append(created_sheet)
        return sheets_created

    @staticmethod
    def _write_insights_sheet(
        workbook: object,
        sheet_name: str,
        insights: list[Insight],
    ) -> str | None:
        """Write one governed insight bucket to a workbook sheet."""

        existing_sheet_name = TemplatePopulationService._find_existing_sheet_name(
            workbook,
            sheet_name,
        )
        if existing_sheet_name is None:
            safe_sheet_name = sanitize_sheet_name(sheet_name, set(workbook.sheetnames))
            worksheet = workbook.create_sheet(safe_sheet_name)
            created_sheet_name = safe_sheet_name
        else:
            worksheet = workbook[existing_sheet_name]
            created_sheet_name = None
            worksheet.delete_rows(1, worksheet.max_row)

        headers = [
            "Year",
            "Source Report Year",
            "Area",
            "Takeaway",
            "Source Section",
            "Page",
            "Confidence",
        ]
        worksheet.append(headers)
        for insight in insights:
            worksheet.append(
                [
                    insight.value_year,
                    insight.source_report_year,
                    insight.area,
                    insight.takeaway,
                    insight.source_section,
                    insight.page_number,
                    insight.confidence,
                ]
            )
        _autosize_columns(worksheet)
        return created_sheet_name

    @staticmethod
    def _is_formula_cell(value: object) -> bool:
        return isinstance(value, str) and value.startswith("=")


def _ordered_unique(values: Iterable[Any]) -> list[Any]:
    seen: set[object] = set()
    ordered: list[object] = []
    for value in values:
        if value in seen:
            continue
        ordered.append(value)
        seen.add(value)
    return ordered


def _duplicate_metrics(metric_values: Iterable[MetricValue]) -> set[str]:
    table_types_by_metric: dict[str, set[str]] = defaultdict(set)
    for metric_value in metric_values:
        table_types_by_metric[metric_value.metric].add(metric_value.table_type)
    return {
        metric
        for metric, table_types in table_types_by_metric.items()
        if len(table_types) > 1
    }


def _row_label(metric: str, table_type: str, duplicate_metrics: set[str]) -> str:
    if metric not in duplicate_metrics:
        return metric
    return f"{metric} ({table_type.replace('_', ' ').title()})"


def _requires_user_decision(sheet_result: SheetValidationResult) -> bool:
    return any(
        "requires user decision" in warning.lower()
        for warning in sheet_result.warnings
    )


def _save_permission_message(output_path: Path) -> str:
    return (
        f"Could not save workbook to '{output_path}'. Close the Excel file if it "
        "is open, verify write permissions, or choose a different output path."
    )


def _autosize_columns(worksheet: Worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = (
            min(max_length + 2, 60)
        )


__all__ = ["TemplatePopulationService"]
