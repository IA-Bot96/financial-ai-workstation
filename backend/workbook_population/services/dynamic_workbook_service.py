"""Dynamic Excel workbook generation service."""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ocr_engine.models.insights_extraction import Insight
from ocr_engine.governance.insight_confidence_governance import (
    InsightConfidenceGovernance,
)
from shared.models.metric_value import MetricValue
from workbook_population.constants.workbook_constants import (
    INSIGHTS_REVIEW_SHEET_NAME,
    INSIGHTS_SHEET_NAME,
)
from workbook_population.models.workbook_cell_mapping import WorkbookCellMappingDraft
from workbook_population.services.sheet_name_sanitizer import sanitize_sheet_name
from workbook_population.services.workbook_mapper import WorkbookMapper


class DynamicWorkbookService:
    """Generate editable financial model workbooks from MetricValue records."""

    def __init__(self, *, mapper: WorkbookMapper | None = None) -> None:
        """Initialize dynamic workbook generation dependencies."""

        self._mapper = mapper or WorkbookMapper()
        self._insight_governance = InsightConfidenceGovernance()
        self._last_cell_mapping_drafts: list[WorkbookCellMappingDraft] = []

    @property
    def last_cell_mapping_drafts(self) -> list[WorkbookCellMappingDraft]:
        """Return cell mappings captured during the most recent generation."""

        return list(self._last_cell_mapping_drafts)

    def generate(
        self,
        *,
        output_file_path: str,
        metric_values: list[MetricValue],
        insights: list[Insight],
    ) -> tuple[list[str], int, list[str]]:
        """Generate a workbook from scratch and save it to output_file_path."""

        self._last_cell_mapping_drafts = []
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        sheet_metrics = self._group_metrics_by_sheet(metric_values)
        created_sheets: list[str] = []
        metrics_written = 0
        existing_sheet_names = set(workbook.sheetnames)

        for sheet_name, values in sheet_metrics.items():
            safe_sheet_name = sanitize_sheet_name(sheet_name, existing_sheet_names)
            worksheet = workbook.create_sheet(safe_sheet_name)
            existing_sheet_names.add(safe_sheet_name)
            created_sheets.append(safe_sheet_name)
            metrics_written += self._write_metric_sheet(worksheet, values)

        governed_insights = self._insight_governance.apply(insights)
        exported_insights = governed_insights.exported_insights
        review_insights = governed_insights.review_insights

        if not sheet_metrics and not exported_insights and not review_insights:
            sheet_name = sanitize_sheet_name("Financial Data", existing_sheet_names)
            worksheet = workbook.create_sheet(sheet_name)
            existing_sheet_names.add(sheet_name)
            worksheet.append(["Metric"])
            _style_header(worksheet[1])
            created_sheets.append(sheet_name)

        if exported_insights:
            sheet_name = self._write_insights_sheet(
                workbook,
                INSIGHTS_SHEET_NAME,
                exported_insights,
            )
            created_sheets.append(sheet_name)
        if review_insights:
            sheet_name = self._write_insights_sheet(
                workbook,
                INSIGHTS_REVIEW_SHEET_NAME,
                review_insights,
            )
            created_sheets.append(sheet_name)

        try:
            output_path = Path(output_file_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            try:
                workbook.save(output_path)
            except PermissionError as exc:
                raise PermissionError(_save_permission_message(output_path)) from exc
        finally:
            workbook.close()

        warnings: list[str] = []
        if not metric_values:
            warnings.append("No metric values were provided; generated an empty model.")

        return created_sheets, metrics_written, warnings

    def _group_metrics_by_sheet(
        self,
        metric_values: Iterable[MetricValue],
    ) -> dict[str, list[MetricValue]]:
        grouped: dict[str, list[MetricValue]] = defaultdict(list)
        for metric_value in metric_values:
            grouped[self._mapper.sheet_name_for_metric_value(metric_value)].append(
                metric_value
            )
        return dict(grouped)

    def _write_metric_sheet(self, worksheet: object, metric_values: list[MetricValue]) -> int:
        years = sorted({metric_value.value_year for metric_value in metric_values})
        row_keys = _ordered_unique(
            (metric_value.metric, metric_value.table_type)
            for metric_value in metric_values
        )
        duplicate_metrics = _duplicate_metrics(metric_values)
        values_by_metric_year = {
            (
                metric_value.metric,
                metric_value.value_year,
                metric_value.table_type,
            ): metric_value.value
            for metric_value in metric_values
        }
        metric_value_by_key = {
            (
                metric_value.metric,
                metric_value.value_year,
                metric_value.table_type,
            ): metric_value
            for metric_value in metric_values
        }

        worksheet.append(["Metric", *years])
        _style_header(worksheet[1])

        written = 0
        for metric, table_type in row_keys:
            row = [_row_label(metric, table_type, duplicate_metrics)]
            for year in years:
                value = values_by_metric_year.get((metric, year, table_type))
                row.append(value)
                if value is not None:
                    written += 1
            worksheet.append(row)
            row_number = worksheet.max_row
            for year_index, year in enumerate(years, start=2):
                metric_value = metric_value_by_key.get((metric, year, table_type))
                if metric_value is None or metric_value.value is None:
                    continue
                self._last_cell_mapping_drafts.append(
                    WorkbookCellMappingDraft(
                        metric=metric_value.metric,
                        value_year=metric_value.value_year,
                        source_report_year=metric_value.source_report_year,
                        table_type=metric_value.table_type,
                        sheet_name=worksheet.title,
                        row=row_number,
                        column=year_index,
                        cell_reference=f"{get_column_letter(year_index)}{row_number}",
                        write_status="written",
                        written_value=metric_value.value,
                    )
                )

        _autosize_columns(worksheet)
        return written

    @staticmethod
    def _write_insights_sheet(
        workbook: Workbook,
        sheet_name: str,
        insights: list[Insight],
    ) -> str:
        safe_sheet_name = sanitize_sheet_name(sheet_name, set(workbook.sheetnames))
        worksheet = workbook.create_sheet(safe_sheet_name)
        worksheet.append(
            [
                "Year",
                "Source Report Year",
                "Area",
                "Takeaway",
                "Source Section",
                "Page",
                "Confidence",
            ]
        )
        _style_header(worksheet[1])

        for insight in insights:
            worksheet.append(
                [
                    insight.value_year,
                    insight.source_report_year,
                    insight.area,
                    insight.takeaway,
                    insight.source_section,
                    insight.page_number,
                    insight.confidence,
                ]
            )

        _autosize_columns(worksheet)
        return safe_sheet_name


def _style_header(cells: object) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in cells:
        cell.fill = fill
        cell.font = font


def _autosize_columns(worksheet: object) -> None:
    for column_cells in worksheet.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = (
            min(max_length + 2, 60)
        )


def _ordered_unique(values: Iterable[Any]) -> list[Any]:
    seen: set[object] = set()
    ordered: list[object] = []
    for value in values:
        if value in seen:
            continue
        ordered.append(value)
        seen.add(value)
    return ordered


def _duplicate_metrics(metric_values: Iterable[MetricValue]) -> set[str]:
    table_types_by_metric: dict[str, set[str]] = defaultdict(set)
    for metric_value in metric_values:
        table_types_by_metric[metric_value.metric].add(metric_value.table_type)
    return {
        metric
        for metric, table_types in table_types_by_metric.items()
        if len(table_types) > 1
    }


def _row_label(metric: str, table_type: str, duplicate_metrics: set[str]) -> str:
    if metric not in duplicate_metrics:
        return metric
    return f"{metric} ({table_type.replace('_', ' ').title()})"


def _save_permission_message(output_path: Path) -> str:
    return (
        f"Could not save workbook to '{output_path}'. Close the Excel file if it "
        "is open, verify write permissions, or choose a different output path."
    )


__all__ = ["DynamicWorkbookService"]
